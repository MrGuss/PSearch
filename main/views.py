from django.shortcuts import render, redirect, get_object_or_404
from .forms import ParaForm, TeacherForm, ReqForm
from .models import Para, Teacher
from django.views.generic.edit import UpdateView, CreateView, DeleteView
from django.views.generic import DetailView, ListView
from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, date
from django.utils.safestring import mark_safe
from django.http import HttpResponseRedirect
from datetime import datetime, timedelta, date
from dal import autocomplete
import re

def createDB(request):
    paras = get_list_of_paras()
    print(len(paras))
    wrong = 0
    teacs = []
    while len(paras)>0:
        para0 = paras[0]
        names = format_names(para0[2][2])
        
        form = ParaForm({'date': para0[0], 'period': para0[1], 'nameC': para0[2][0],
                         'categ': para0[2][1], 'nameT': '\n'.join(names), 'place':para0[2][3]})
        if not(form.is_valid()):
            wrong+=1
        else:
            form.save()
            for name in names:
                if not(name in teacs):
                    teacs.append(name)
                    #print(len(name), type(name), name)
                    formT = TeacherForm({'nameT': name})
                    formT.save()
                
            #print({'date': para0[0], 'period': para0[1], 'nameC': para0[2][0],
            #            'categ': para0[2][1], 'nameT':para0[2][2], 'place':para0[2][3]})
        paras.pop(0)
    return redirect('/search')

def home(request):
    days = {0:"Понедельник", 1:"Вторник", 2:"Среда", 3:"Четверг", 4:"Пятница", 5:"Суббота"}
    response = []
    if request.method == 'POST':
        form = ReqForm(request.POST)
        nedelya = int(request.POST['nedelya'])
        print(request.POST['ID'])
        name = Teacher.objects.filter(Id=int(request.POST['ID']))[0].nameT
        pari = Para.objects.filter(nameT__contains=name)
        for par in pari:
            if (nedelya%2)==((par.period+1)%2):
                response.append([days[par.date], par.period//2+1, par.nameC, par.categ, par.nameT, par.place])
        print(response)
    else:
        form = ReqForm()
    context = {'form':form, 'resp':response}
    return render(request, "main/search.html", context)

class TeacherAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Teacher.objects.all()
        if self.q:
            qs = qs.filter(nameT__icontains=self.q)
        return qs






def format_names(names_str):
    # Define the regular expression pattern to match names
    name_pattern = r'[А-Яа-яЁё]+ [А-Яа-яЁё]\.[А-Яа-яЁё]\.'

    # Extract all names from the string
    names = re.findall(name_pattern, names_str)
    # Format each name accordingly
    return names

def test_substrings(test_str, substr_list):
    for substr in substr_list:
        if substr in test_str:
            return False
    return True

def get_list_of_paras():
    from bs4 import BeautifulSoup
    import requests
    from openpyxl import load_workbook
    from time import sleep
    def show(arr):
        for i in arr:
            print(i)

    url = 'https://www.mirea.ru/schedule/'

    response = requests.get(url)
    html = response.content

    soup = BeautifulSoup(html, 'html.parser')
    tags = soup.find_all(class_='uk-link-toggle')

    hrefs = [tag.get('href') for tag in tags][:-3]
    urls = []
    for href in hrefs:
        if not(('ekz' in href[href.rfind("/"):]) | ('sessiya' in href[href.rfind("/"):])):
            urls.append(href)
    pari = []

    for url in urls:
        #print(url)
        response = requests.get(url)
        #print(url[url.rfind("/")+1:])
        with open("xlsx/"+url[url.rfind("/")+1:], 'wb') as file:
            file.write(response.content)
            #print("dsdfsfdfsdfsdf")
        #print(response.content)
        wb = load_workbook(filename = "xlsx/"+url[url.rfind("/")+1:], read_only=True)
        ws = wb.active
        i = 0
        table = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            #print(i, len(row), row)
            table.append(row)
            i+=1
            if i==84:
                break
        week = [table[i*14:i*14+13] for i in range(6)]
        dw = 0
        for day in week:
            time = 0
            for para in day:
                predmets = [para[i:i+15] for i in range(0, len(para), 15)]
                for i in predmets:
                    if len(i)>=12:
                        if (i[7]!='') & (i[7]!=None):
                            if not([dw, time, i[5:9]] in pari):
                                pari.append([dw, time, i[5:9]])
                                #print(dw, time, i[5:9])
                        if (i[12]!='') & (i[12]!=None):
                            if not([dw, time, i[10:14]] in pari):
                                pari.append([dw, time, i[10:14]])
                                #print(dw, time, i[10:14])
                #sleep(2)
                time+=1
            dw+=1
    #abcdeFghijKlmnopqrstUvwxqZaaabacadaeafagahaiAJ
    #0-83
    return pari
